"""Flujo de validacion de actividades externas.

- El becario (creador) completa los datos y envia la actividad a revision
  (status=pending_validation).
- Un administrador la valida (status=validated) o la rechaza (vuelve a active).
- La validacion registra fecha (validated_at) y responsable (validated_by).
- El administrador puede listar y exportar las actividades externas.
"""

from datetime import UTC, datetime, timedelta

from app.db.constants import MVP_TENANT_ID
from app.db.enums import ActivityStatus
from app.db.models.activities import Activity
from factories import auth_cookies, auth_headers, make_user

PNG = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+M8AAAMBAQDJ/pLvAAAAAElFTkSuQmCC"


def _make_external_activity(db, owner, *, relevant_data=None, hours=None):
    act = Activity(
        title="Voluntariado externo",
        zone="Caracas",
        raw_address="Av. Principal",
        date_time=datetime.now(UTC) - timedelta(days=1),
        creator_id=owner.id,
        status=ActivityStatus.active.value,
        tenant_id=MVP_TENANT_ID,
        external_beneficiary="Fundacion Ejemplo",
        external_supervisor="Ana Perez",
        external_supervisor_email="ana@ejemplo.org",
        external_assigned_hours=hours,
        external_relevant_data=relevant_data,
    )
    db.add(act)
    db.commit()
    db.refresh(act)
    return act


def _upload_evidence(client, db, act, user):
    resp = client.post(
        f"/api/v1/activities/{act.id}/evidence",
        json={"images": [PNG]},
        cookies=auth_cookies(user),
        headers=auth_headers(),
    )
    assert resp.status_code == 200, resp.text


def test_submit_requires_relevant_data(client, db):
    owner = make_user(db, auth_source="google", status="active")
    act = _make_external_activity(db, owner, relevant_data=None, hours=4)
    _upload_evidence(client, db, act, owner)

    resp = client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )
    assert resp.status_code == 422
    assert resp.json()["error"]["code"] == "validation.missing_field"


def test_submit_requires_hours(client, db):
    owner = make_user(db, auth_source="google", status="active")
    act = _make_external_activity(db, owner, relevant_data="Apoyo en comedor", hours=None)
    _upload_evidence(client, db, act, owner)

    resp = client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )
    assert resp.status_code == 422


def test_submit_requires_evidence(client, db):
    owner = make_user(db, auth_source="google", status="active")
    act = _make_external_activity(db, owner, relevant_data="Apoyo en comedor", hours=4)

    resp = client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )
    assert resp.status_code == 422


def test_full_validation_flow(client, db):
    owner = make_user(db, auth_source="google", status="active")
    admin = make_user(db, role="admin", status="active")
    act = _make_external_activity(db, owner, relevant_data="Apoyo en comedor", hours=4)
    _upload_evidence(client, db, act, owner)

    # Becario envia a validacion.
    resp = client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )
    assert resp.status_code == 200
    assert resp.json()["status"] == "pending_validation"

    # No-admin no puede validar.
    other = make_user(db, auth_source="google", status="active")
    resp = client.post(
        f"/api/v1/activities/{act.id}/validate",
        json={"notes": "ok"},
        cookies=auth_cookies(other),
        headers=auth_headers(),
    )
    assert resp.status_code == 403

    # Admin valida.
    resp = client.post(
        f"/api/v1/activities/{act.id}/validate",
        json={"notes": "Valido"},
        cookies=auth_cookies(admin),
        headers=auth_headers(),
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "validated"
    assert body["validated_by"] == str(admin.id)
    assert body["validated_by_name"] == admin.name
    assert body["validated_at"] is not None
    assert body["validation_notes"] == "Valido"


def test_reject_returns_to_active(client, db):
    owner = make_user(db, auth_source="google", status="active")
    admin = make_user(db, role="admin", status="active")
    act = _make_external_activity(db, owner, relevant_data="Apoyo en comedor", hours=4)
    _upload_evidence(client, db, act, owner)
    client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )

    resp = client.post(
        f"/api/v1/activities/{act.id}/reject-validation",
        json={"notes": "Faltan firmas"},
        cookies=auth_cookies(admin),
        headers=auth_headers(),
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "active"
    assert body["validated_by"] is None
    assert body["validation_notes"] == "Faltan firmas"

    # Rechazo sin motivo es invalido.
    resp = client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )
    assert resp.status_code == 200
    resp = client.post(
        f"/api/v1/activities/{act.id}/reject-validation",
        json={"notes": ""},
        cookies=auth_cookies(admin),
        headers=auth_headers(),
    )
    assert resp.status_code == 422


def test_admin_list_external_validation(client, db):
    owner = make_user(db, auth_source="google", status="active")
    admin = make_user(db, role="admin", status="active")
    act = _make_external_activity(db, owner, relevant_data="Apoyo", hours=4)
    _upload_evidence(client, db, act, owner)
    client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )

    # Usuario normal no accede.
    other = make_user(db, auth_source="google", status="active")
    resp = client.get(
        "/api/v1/activities/admin/external-validation",
        cookies=auth_cookies(other),
        headers=auth_headers(),
    )
    assert resp.status_code == 403

    resp = client.get(
        "/api/v1/activities/admin/external-validation",
        cookies=auth_cookies(admin),
        headers=auth_headers(),
    )
    assert resp.status_code == 200
    assert any(a["id"] == str(act.id) for a in resp.json())


def test_export_external_activities_returns_zip(client, db):
    import io
    import zipfile

    owner = make_user(db, auth_source="google", status="active")
    admin = make_user(db, role="admin", status="active")
    act = _make_external_activity(db, owner, relevant_data="Apoyo", hours=4)
    _upload_evidence(client, db, act, owner)
    client.post(
        f"/api/v1/activities/{act.id}/submit-validation",
        cookies=auth_cookies(owner),
        headers=auth_headers(),
    )
    client.post(
        f"/api/v1/activities/{act.id}/validate",
        json={"notes": "ok"},
        cookies=auth_cookies(admin),
        headers=auth_headers(),
    )

    resp = client.get(
        "/api/v1/activities/admin/export-external?status=validated",
        cookies=auth_cookies(admin),
        headers=auth_headers(),
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/zip"

    zf = zipfile.ZipFile(io.BytesIO(resp.content))
    names = zf.namelist()
    assert "actividades_externas.xlsx" in names
    assert any(n.startswith(f"adjuntos/{act.id}/evidencia/") for n in names)

    # El Excel contiene la fila de la actividad y la ruta relativa de adjuntos.
    from openpyxl import load_workbook

    xlsx_bytes = zf.read("actividades_externas.xlsx")
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [c for c in rows[0]]
    assert "Ruta adjuntos" in header
    col = header.index("Ruta adjuntos")
    assert any(r[col] == f"adjuntos/{act.id}/" for r in rows[1:])
